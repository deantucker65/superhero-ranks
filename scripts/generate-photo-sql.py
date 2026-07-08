#!/usr/bin/env python3
"""Read the content spreadsheet and emit SQL UPDATE statements to set photo_url
for actors and characters. Skips empty and fragile (Google/gstatic thumbnail)
URLs so only stable links reach production.

Usage:
  python3 scripts/generate-photo-sql.py [path/to/workbook.xlsx]
Output: writes content/photo-updates.sql (paste it into the Supabase SQL Editor).
"""
import sys
from openpyxl import load_workbook

SRC = sys.argv[1] if len(sys.argv) > 1 else 'content/superhero-content.xlsx'
OUT = 'content/photo-updates.sql'

def fragile(u):
    s = str(u).lower()
    return ('gstatic.com' in s) or ('images?q=tbn' in s) or ('lookaside' in s) or ('google.com/imgres' in s)

def q(v):
    return "'" + str(v).replace("'", "''") + "'"

wb = load_workbook(SRC, data_only=True)
lines, n_actor, n_char, skipped = [], 0, 0, 0

a = wb['Actors']
for r in range(2, a.max_row + 1):
    name, url = a.cell(r, 1).value, a.cell(r, 2).value
    if not name or not url:
        continue
    if fragile(url):
        skipped += 1
        continue
    lines.append(f"update public.actors set photo_url = {q(url)} where name = {q(name)};")
    n_actor += 1

lines.append("")

c = wb['Characters']
for r in range(2, c.max_row + 1):
    actor, char, url = c.cell(r, 1).value, c.cell(r, 2).value, c.cell(r, 8).value
    if not actor or not char or not url:
        continue
    if fragile(url):
        skipped += 1
        continue
    lines.append(
        f"update public.characters set photo_url = {q(url)} "
        f"where name = {q(char)} and actor_id = (select id from public.actors where name = {q(actor)});"
    )
    n_char += 1

header = [
    f"-- Photo URL updates generated from {SRC}",
    f"-- {n_actor} actor(s), {n_char} character(s). Skipped {skipped} fragile/thumbnail link(s).",
    "-- Review, then run in the Supabase SQL Editor.",
    "begin;",
]
body = "\n".join(header + [""] + lines + ["", "commit;", ""])
open(OUT, 'w').write(body)
print(f"Wrote {OUT}: {n_actor} actor + {n_char} character updates ({skipped} fragile skipped).")
